"""
QR Code & Barcode Scanner - Android App
Dung Kivy Camera widget + zxingcpp de decode QR/Barcode.
Build: buildozer android debug
"""
 
import datetime
import os
import threading
 
from kivy.app import App
from kivy.lang import Builder
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.uix.camera import Camera
from kivy.clock import Clock
from kivy.utils import platform
from kivy.graphics.texture import Texture
 
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False
 
try:
    import zxingcpp
    ZXING_OK = True
except ImportError:
    ZXING_OK = False
 
KV = """
#:import MDApp kivymd.app.MDApp
 
ScreenManager:
    id: sm
    ScanScreen:
        name: "scan"
    ListScreen:
        name: "list"
 
<ScanScreen>:
    BoxLayout:
        orientation: "vertical"
        canvas.before:
            Color:
                rgba: 0.96, 0.96, 0.96, 1
            Rectangle:
                pos: self.pos
                size: self.size
 
        # Top bar
        BoxLayout:
            size_hint_y: None
            height: "56dp"
            padding: "8dp"
            spacing: "8dp"
            canvas.before:
                Color:
                    rgba: 0.098, 0.463, 0.824, 1
                Rectangle:
                    pos: self.pos
                    size: self.size
 
            Label:
                text: "QR & Barcode Scanner"
                font_size: "18sp"
                bold: True
                color: 1, 1, 1, 1
 
            Button:
                text: "Danh sach"
                size_hint_x: None
                width: "110dp"
                background_color: 0.263, 0.627, 0.278, 1
                on_release: app.go_list()
 
        # Camera
        Camera:
            id: camera
            resolution: (640, 480)
            play: True
 
        # Bottom bar
        BoxLayout:
            size_hint_y: None
            height: "52dp"
            padding: "8dp", "4dp"
            spacing: "8dp"
            canvas.before:
                Color:
                    rgba: 1, 1, 1, 1
                Rectangle:
                    pos: self.pos
                    size: self.size
 
            Label:
                id: scan_status
                text: "Chua quet ma nao"
                color: 0.46, 0.46, 0.46, 1
                font_size: "14sp"
 
<ListScreen>:
    BoxLayout:
        orientation: "vertical"
 
        # Top bar
        BoxLayout:
            size_hint_y: None
            height: "56dp"
            padding: "8dp"
            spacing: "8dp"
            canvas.before:
                Color:
                    rgba: 0.098, 0.463, 0.824, 1
                Rectangle:
                    pos: self.pos
                    size: self.size
 
            Button:
                text: "<  Quay lai"
                size_hint_x: None
                width: "100dp"
                background_color: 0, 0, 0, 0
                color: 1, 1, 1, 1
                on_release: app.go_scan()
 
            Label:
                text: "Danh sach QR"
                font_size: "18sp"
                bold: True
                color: 1, 1, 1, 1
 
            Button:
                text: "Xuat Excel"
                size_hint_x: None
                width: "110dp"
                background_color: 0.263, 0.627, 0.278, 1
                on_release: app.export_excel()
 
            Button:
                text: "Xoa het"
                size_hint_x: None
                width: "80dp"
                background_color: 0.898, 0.224, 0.208, 1
                on_release: app.clear_all()
 
        # List
        ScrollView:
            GridLayout:
                id: qr_list
                cols: 1
                size_hint_y: None
                height: self.minimum_height
                spacing: "2dp"
                padding: "4dp"
 
        # Footer count
        BoxLayout:
            size_hint_y: None
            height: "36dp"
            padding: "12dp", "4dp"
            canvas.before:
                Color:
                    rgba: 0.94, 0.94, 0.94, 1
                Rectangle:
                    pos: self.pos
                    size: self.size
            Label:
                id: count_label
                text: "0 ma QR"
                color: 0.46, 0.46, 0.46, 1
                font_size: "13sp"
                halign: "left"
                text_size: self.size
 
<QRItem>:
    size_hint_y: None
    height: "56dp"
    padding: "12dp", "4dp"
    spacing: "8dp"
    canvas.before:
        Color:
            rgba: root.bg_color
        Rectangle:
            pos: self.pos
            size: self.size
    Label:
        text: root.idx_text
        size_hint_x: None
        width: "40dp"
        color: 0.098, 0.463, 0.824, 1
        bold: True
        font_size: "13sp"
    Label:
        text: root.qr_text
        color: 0.13, 0.13, 0.13, 1
        font_size: "12sp"
        halign: "left"
        text_size: self.size
    Label:
        text: root.time_text
        size_hint_x: None
        width: "150dp"
        color: 0.46, 0.46, 0.46, 1
        font_size: "11sp"
        halign: "right"
        text_size: self.size
"""
 
 
class ScanScreen(Screen):
    pass
 
 
class ListScreen(Screen):
    pass
 
 
# Simple list item widget
from kivy.uix.boxlayout import BoxLayout
from kivy.properties import StringProperty, ListProperty
 
 
class QRItem(BoxLayout):
    idx_text  = StringProperty("")
    qr_text   = StringProperty("")
    time_text = StringProperty("")
    bg_color  = ListProperty([1, 1, 1, 1])
 
 
class QRScanApp(App):
    def build(self):
        self.scan_data  = []
        self.seen_qrs   = set()
        self._scanning  = False
        self._dialog    = None
        root = Builder.load_string(KV)
        # Start scan loop
        Clock.schedule_interval(self._scan_frame, 1.0 / 10)  # 10 fps
        return root
 
    def go_list(self):
        self.root.current = "list"
 
    def go_scan(self):
        self.root.current = "scan"
 
    # ── Scan loop ─────────────────────────────────────────────────────────────
    def _scan_frame(self, dt):
        if not ZXING_OK:
            return
        try:
            cam = self.root.get_screen("scan").ids.camera
            if not cam.play:
                return
            tex = cam.texture
            if tex is None:
                return
            # Convert texture to bytes and decode
            import numpy as np
            from PIL import Image as PILImage
            buf = tex.pixels
            img_arr = np.frombuffer(buf, dtype=np.uint8).reshape(
                tex.height, tex.width, 4)
            # zxingcpp expects RGB
            rgb = img_arr[:, :, :3]
            results = zxingcpp.read_barcodes(rgb)
            for r in results:
                self._register_code(r.text, r.format.name)
        except Exception:
            pass
 
    # ── Data ──────────────────────────────────────────────────────────────────
    def _register_code(self, data: str, code_type: str = "QRCODE"):
        if not data or data in self.seen_qrs:
            return
        self.seen_qrs.add(data)
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        idx = len(self.scan_data) + 1
        self.scan_data.append({"index": idx, "type": code_type,
                                "qr": data, "time": now})
 
        # Update scan status
        lbl = self.root.get_screen("scan").ids.scan_status
        lbl.text = f"Da quet {idx} ma"
 
        # Add row to list
        item = QRItem(
            idx_text  = str(idx),
            qr_text   = data[:55] + ("..." if len(data) > 55 else ""),
            time_text = now,
            bg_color  = [0.89, 0.95, 1, 1] if idx % 2 == 0 else [1, 1, 1, 1])
 
        self.root.get_screen("list").ids.qr_list.add_widget(item)
        self.root.get_screen("list").ids.count_label.text = f"{idx} ma"
 
    def clear_all(self):
        self.scan_data.clear()
        self.seen_qrs.clear()
        self.root.get_screen("list").ids.qr_list.clear_widgets()
        self.root.get_screen("list").ids.count_label.text = "0 ma"
        self.root.get_screen("scan").ids.scan_status.text = "Chua quet ma nao"
 
    # ── Excel export ──────────────────────────────────────────────────────────
    def export_excel(self):
        if not OPENPYXL_OK:
            return
        if not self.scan_data:
            return
        threading.Thread(target=self._write_excel, daemon=True).start()
 
    def _write_excel(self):
        if platform == "android":
            from android.storage import primary_external_storage_path
            folder = os.path.join(primary_external_storage_path(), "Download")
        else:
            folder = os.path.expanduser("~")
        os.makedirs(folder, exist_ok=True)
 
        fname = f"QR_Scan_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        path  = os.path.join(folder, fname)
 
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "QR Scan Data"
 
            thin   = Side(style="thin", color="B0BEC5")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            hdr_fill  = PatternFill("solid", fgColor="1976D2")
            hdr_font  = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
            hdr_align = Alignment(horizontal="center", vertical="center")
            even_fill = PatternFill("solid", fgColor="E3F2FD")
 
            # Title
            ws.merge_cells("A1:C1")
            tc = ws["A1"]
            tc.value     = "QR Code & Barcode Scan Report"
            tc.font      = Font(bold=True, size=14, color="1565C0", name="Calibri")
            tc.alignment = Alignment(horizontal="center")
            ws.row_dimensions[1].height = 28
            ws["A2"] = "Ngay xuat:"
            ws["B2"] = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            ws["A3"] = "Tong so ma:"
            ws["B3"] = len(self.scan_data)
 
            # Header row 5: STT | QRCode | Time
            for col, h in enumerate(["STT", "QRCode", "Time"], 1):
                cell = ws.cell(row=5, column=col, value=h)
                cell.font      = hdr_font
                cell.fill      = hdr_fill
                cell.alignment = hdr_align
                cell.border    = border
            ws.row_dimensions[5].height = 22
 
            # Data
            for row_idx, d in enumerate(self.scan_data, 6):
                fill = even_fill if row_idx % 2 == 0 else PatternFill()
                for c, v in [(1, d["index"]), (2, d["qr"]), (3, d["time"])]:
                    cell = ws.cell(row=row_idx, column=c, value=v)
                    cell.fill      = fill
                    cell.border    = border
                    cell.font      = Font(name="Calibri", size=10)
                    cell.alignment = Alignment(
                        vertical="center",
                        horizontal="center" if c in (1, 3) else "left",
                        wrap_text=(c == 2))
 
            ws.column_dimensions["A"].width = 8
            ws.column_dimensions["B"].width = 60
            ws.column_dimensions["C"].width = 22
            ws.freeze_panes = "A6"
            wb.save(path)
            Clock.schedule_once(
                lambda _: self._show_toast(f"Da luu: {fname}"), 0)
        except Exception as ex:
            Clock.schedule_once(
                lambda _: self._show_toast(f"Loi: {ex}"), 0)
 
    def _show_toast(self, msg):
        # Simple toast using a popup
        from kivy.uix.popup import Popup
        from kivy.uix.label import Label
        p = Popup(title="", content=Label(text=msg),
                  size_hint=(0.8, 0.2), auto_dismiss=True)
        p.open()
        Clock.schedule_once(lambda _: p.dismiss(), 3)
 
    def on_stop(self):
        pass
 
 
if __name__ == "__main__":
    QRScanApp().run()
